"""
Модуль экспорта данных в Excel.
Задача 10: ExcelExporter
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime, date
from collections import defaultdict

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

try:
    import pandas as pd
    HAVE_PANDAS = True
except ImportError:
    HAVE_PANDAS = False

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Экспортер данных актов взвешивания в Excel.
    
    Функции:
    - Группировка актов по дате
    - Суммирование показателей за день
    - Создание Excel файла с форматированием
    - Применение стилей
    """
    
    def __init__(self):
        if not HAVE_OPENPYXL and not HAVE_PANDAS:
            raise ImportError(
                "Требуется openpyxl или pandas. "
                "Установите: pip install openpyxl pandas"
            )
        
        self.workbook = None
        self.use_openpyxl = HAVE_OPENPYXL
        
        logger.info("ExcelExporter инициализирован")
    
    def group_acts_by_date(
        self,
        acts: List[Dict[str, Any]]
    ) -> Dict[date, List[Dict[str, Any]]]:
        """
        Группирует акты по дате.
        
        Args:
            acts: список актов взвешивания
            
        Returns:
            Словарь {дата: [акты]}
        """
        grouped = defaultdict(list)
        
        for act in acts:
            # Извлекаем дату
            act_date = None
            if 'started_at' in act:
                if isinstance(act['started_at'], str):
                    act_date = datetime.fromisoformat(act['started_at']).date()
                elif isinstance(act['started_at'], datetime):
                    act_date = act['started_at'].date()
            elif 'date' in act:
                if isinstance(act['date'], str):
                    act_date = datetime.fromisoformat(act['date']).date()
                elif isinstance(act['date'], (datetime, date)):
                    act_date = act['date'] if isinstance(act['date'], date) else act['date'].date()
            
            if act_date:
                grouped[act_date].append(act)
        
        logger.info(f"Акты сгруппированы: {len(grouped)} дней, {len(acts)} актов")
        return dict(grouped)
    
    def summarize_by_date(
        self,
        grouped_acts: Dict[date, List[Dict[str, Any]]]
    ) -> List[Dict[str, Any]]:
        """
        Суммирует показатели по дням.
        
        Returns:
            Список сводок по дням
        """
        summaries = []
        
        for act_date, acts in sorted(grouped_acts.items()):
            summary = {
                'date': act_date,
                'acts_count': len(acts),
                'total_left': sum(act.get('left_count', 0) for act in acts),
                'total_right': sum(act.get('right_count', 0) for act in acts),
                'total_crossings': sum(
                    act.get('left_count', 0) + act.get('right_count', 0)
                    for act in acts
                ),
                'avg_duration': sum(act.get('duration', 0) for act in acts) / len(acts) if acts else 0,
                'max_peak': max((act.get('peak_count', 0) for act in acts), default=0),
                'total_seen': sum(act.get('seen_total', 0) for act in acts)
            }
            summaries.append(summary)
        
        logger.info(f"Создано сводок: {len(summaries)}")
        return summaries
    
    def export_to_excel(
        self,
        acts: List[Dict[str, Any]],
        output_path: str,
        group_by_date: bool = True
    ) -> bool:
        """
        Экспортирует акты в Excel файл.
        
        Args:
            acts: список актов
            output_path: путь к выходному файлу
            group_by_date: группировать ли по датам
            
        Returns:
            True если успешно
        """
        try:
            if self.use_openpyxl:
                return self._export_openpyxl(acts, output_path, group_by_date)
            else:
                return self._export_pandas(acts, output_path, group_by_date)
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта: {e}")
            return False
    
    def _export_openpyxl(
        self,
        acts: List[Dict[str, Any]],
        output_path: str,
        group_by_date: bool
    ) -> bool:
        """Экспорт через openpyxl с форматированием"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Акты взвешивания"
        
        # Стили
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if group_by_date:
            # Группировка по датам
            grouped = self.group_acts_by_date(acts)
            summaries = self.summarize_by_date(grouped)
            
            # Заголовки
            headers = ['Дата', 'Актов', 'Вход слева', 'Вход справа', 'Всего', 'Ср. длительность', 'Пик', 'Уникальных']
            ws.append(headers)
            
            # Применяем стили к заголовкам
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Данные
            for summary in summaries:
                ws.append([
                    summary['date'].strftime('%Y-%m-%d'),
                    summary['acts_count'],
                    summary['total_left'],
                    summary['total_right'],
                    summary['total_crossings'],
                    f"{summary['avg_duration']:.1f}",
                    summary['max_peak'],
                    summary['total_seen']
                ])
        else:
            # Все акты
            headers = ['ID', 'Начало', 'Окончание', 'Длительность', 'Слева', 'Справа', 'Пик', 'Уникальных']
            ws.append(headers)
            
            # Стили заголовков
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Данные
            for act in acts:
                ws.append([
                    act.get('id', act.get('act_id', '')),
                    act.get('started_at', ''),
                    act.get('ended_at', ''),
                    f"{act.get('duration', 0):.1f}",
                    act.get('left_count', 0),
                    act.get('right_count', 0),
                    act.get('peak_count', 0),
                    act.get('seen_total', 0)
                ])
        
        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение
        wb.save(output_path)
        logger.info(f"✅ Excel файл создан: {output_path}")
        return True
    
    def _export_pandas(
        self,
        acts: List[Dict[str, Any]],
        output_path: str,
        group_by_date: bool
    ) -> bool:
        """Экспорт через pandas"""
        if group_by_date:
            grouped = self.group_acts_by_date(acts)
            summaries = self.summarize_by_date(grouped)
            df = pd.DataFrame(summaries)
        else:
            df = pd.DataFrame(acts)
        
        # Экспорт
        df.to_excel(output_path, index=False, engine='openpyxl')
        logger.info(f"✅ Excel файл создан: {output_path}")
        return True
