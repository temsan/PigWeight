"""
Модуль анализа Excel файлов с ручными записями операторов.
Задача 9: ExcelAnalyzer
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import json

try:
    import openpyxl
    from openpyxl import load_workbook
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

try:
    import pandas as pd
    HAVE_PANDAS = True
except ImportError:
    HAVE_PANDAS = False

logger = logging.getLogger(__name__)


class ExcelAnalyzer:
    """
    Анализатор Excel файлов с ручными записями взвешиваний.
    
    Функции:
    - Парсинг Excel файла
    - Определение структуры столбцов
    - Извлечение данных: секция, дата, пары (вес, количество), итоги
    - Анализ схемы шаблона
    """
    
    def __init__(self, excel_path: str):
        if not HAVE_OPENPYXL and not HAVE_PANDAS:
            raise ImportError(
                "Требуется openpyxl или pandas. "
                "Установите: pip install openpyxl pandas"
            )
        
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel файл не найден: {excel_path}")
        
        self.workbook = None
        self.schema = None
        self.data = []
        
        logger.info(f"ExcelAnalyzer инициализирован: {excel_path}")
    
    def load(self) -> bool:
        """Загружает Excel файл"""
        try:
            if HAVE_OPENPYXL:
                self.workbook = load_workbook(self.excel_path, data_only=True)
                logger.info(f"✅ Excel загружен через openpyxl")
                return True
            elif HAVE_PANDAS:
                self.workbook = pd.ExcelFile(self.excel_path)
                logger.info(f"✅ Excel загружен через pandas")
                return True
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки Excel: {e}")
            return False
    
    def analyze_schema(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Анализирует схему Excel файла.
        
        Returns:
            Словарь со схемой: позиции столбцов, форматы, структура
        """
        if not self.workbook:
            self.load()
        
        schema = {
            'file_path': str(self.excel_path),
            'sheets': [],
            'columns': {},
            'date_format': None,
            'number_format': None
        }
        
        try:
            if HAVE_OPENPYXL:
                # Анализ через openpyxl
                sheet = self.workbook.active if not sheet_name else self.workbook[sheet_name]
                
                # Получаем заголовки (первая строка)
                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append({
                            'column': cell.column_letter,
                            'index': cell.column,
                            'name': str(cell.value),
                            'type': type(cell.value).__name__
                        })
                
                schema['columns'] = headers
                schema['sheets'] = [s for s in self.workbook.sheetnames]
                
                # Определяем форматы
                for row in sheet.iter_rows(min_row=2, max_row=10):
                    for cell in row:
                        if cell.value and isinstance(cell.value, datetime):
                            schema['date_format'] = cell.number_format
                            break
                
            elif HAVE_PANDAS:
                # Анализ через pandas
                df = pd.read_excel(self.excel_path, sheet_name=sheet_name or 0)
                
                schema['columns'] = [
                    {
                        'index': i,
                        'name': col,
                        'type': str(df[col].dtype)
                    }
                    for i, col in enumerate(df.columns)
                ]
                schema['sheets'] = self.workbook.sheet_names
            
            self.schema = schema
            logger.info(f"✅ Схема проанализирована: {len(schema['columns'])} столбцов")
            return schema
            
        except Exception as e:
            logger.error(f"❌ Ошибка анализа схемы: {e}")
            return schema
    
    def parse_data(self, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Парсит данные из Excel файла.
        
        Returns:
            Список записей с данными: секция, дата, вес, количество, итоги
        """
        if not self.workbook:
            self.load()
        
        records = []
        
        try:
            if HAVE_PANDAS:
                # Парсинг через pandas (проще и быстрее)
                df = pd.read_excel(self.excel_path, sheet_name=sheet_name or 0)
                
                for idx, row in df.iterrows():
                    record = {
                        'row_number': idx + 2,  # +2 т.к. индекс с 0 и есть заголовок
                        'data': row.to_dict()
                    }
                    
                    # Пытаемся извлечь стандартные поля
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if 'дата' in col_lower or 'date' in col_lower:
                            record['date'] = row[col]
                        elif 'секц' in col_lower or 'section' in col_lower:
                            record['section'] = row[col]
                        elif 'вес' in col_lower or 'weight' in col_lower:
                            record['weight'] = row[col]
                        elif 'количество' in col_lower or 'count' in col_lower or 'кол' in col_lower:
                            record['count'] = row[col]
                    
                    records.append(record)
                
            elif HAVE_OPENPYXL:
                # Парсинг через openpyxl
                sheet = self.workbook.active if not sheet_name else self.workbook[sheet_name]
                
                # Получаем заголовки
                headers = [cell.value for cell in sheet[1]]
                
                # Читаем данные
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Пропускаем пустые строки
                        continue
                    
                    record = {
                        'row_number': row_idx,
                        'data': dict(zip(headers, row))
                    }
                    
                    # Извлекаем стандартные поля
                    for i, header in enumerate(headers):
                        if not header:
                            continue
                        header_lower = str(header).lower()
                        value = row[i]
                        
                        if 'дата' in header_lower or 'date' in header_lower:
                            record['date'] = value
                        elif 'секц' in header_lower or 'section' in header_lower:
                            record['section'] = value
                        elif 'вес' in header_lower or 'weight' in header_lower:
                            record['weight'] = value
                        elif 'количество' in header_lower or 'count' in header_lower or 'кол' in header_lower:
                            record['count'] = value
                    
                    records.append(record)
            
            self.data = records
            logger.info(f"✅ Данные извлечены: {len(records)} записей")
            return records
            
        except Exception as e:
            logger.error(f"❌ Ошибка парсинга данных: {e}")
            return records
    
    def save_schema(self, output_path: str) -> bool:
        """Сохраняет схему в JSON файл"""
        if not self.schema:
            self.analyze_schema()
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.schema, f, indent=2, ensure_ascii=False, default=str)
            logger.info(f"✅ Схема сохранена: {output_path}")
            return True
        except Exception as e:
            logger.error(f"❌ Ошибка сохранения схемы: {e}")
            return False
    
    def get_summary(self) -> Dict[str, Any]:
        """Возвращает сводку по файлу"""
        if not self.data:
            self.parse_data()
        
        summary = {
            'file': str(self.excel_path),
            'total_records': len(self.data),
            'sheets': len(self.schema['sheets']) if self.schema else 0,
            'columns': len(self.schema['columns']) if self.schema else 0
        }
        
        # Статистика по датам
        dates = [r.get('date') for r in self.data if r.get('date')]
        if dates:
            summary['date_range'] = {
                'from': min(dates),
                'to': max(dates),
                'unique_dates': len(set(dates))
            }
        
        # Статистика по секциям
        sections = [r.get('section') for r in self.data if r.get('section')]
        if sections:
            summary['sections'] = list(set(sections))
        
        return summary
    
    def close(self):
        """Закрывает файл"""
        if self.workbook and HAVE_OPENPYXL:
            self.workbook.close()
        self.workbook = None
