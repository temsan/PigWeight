"""
Модуль сверки автоматических актов с ручными записями из Excel.
Задача 11: ExcelComparator
"""

import logging
from typing import Dict, List, Any, Tuple, Optional
from datetime import datetime, timedelta
from dataclasses import dataclass
import math

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

logger = logging.getLogger(__name__)


@dataclass
class ComparisonResult:
    """Результат сравнения одного акта"""
    auto_act: Dict[str, Any]
    manual_act: Optional[Dict[str, Any]]
    match_type: str  # 'exact', 'close', 'mismatch', 'missing'
    time_diff_minutes: float
    count_diff_percent: float
    weight_diff_percent: float = 0.0


class ExcelComparator:
    """
    Сверка автоматических актов с ручными записями.
    
    Функции:
    - Сопоставление актов по времени (±5 минут)
    - Сравнение показателей
    - Вычисление метрик точности
    - Генерация отчета о сверке
    """
    
    def __init__(
        self,
        time_tolerance_minutes: float = 5.0,
        count_tolerance_percent: float = 10.0
    ):
        self.time_tolerance = timedelta(minutes=time_tolerance_minutes)
        self.count_tolerance = count_tolerance_percent
        
        self.comparisons: List[ComparisonResult] = []
        self.metrics = {}
        
        logger.info(
            f"ExcelComparator инициализирован: "
            f"time_tolerance={time_tolerance_minutes}мин, "
            f"count_tolerance={count_tolerance_percent}%"
        )
    
    def match_acts_by_time(
        self,
        auto_acts: List[Dict[str, Any]],
        manual_acts: List[Dict[str, Any]]
    ) -> List[ComparisonResult]:
        """
        Сопоставляет автоматические и ручные акты по времени.
        
        Args:
            auto_acts: автоматические акты из БД
            manual_acts: ручные записи из Excel
            
        Returns:
            Список результатов сравнения
        """
        results = []
        matched_manual = set()
        
        for auto_act in auto_acts:
            # Извлекаем время автоматического акта
            auto_time = self._extract_datetime(auto_act.get('started_at'))
            if not auto_time:
                continue
            
            # Ищем ближайший ручной акт
            best_match = None
            min_diff = None
            
            for idx, manual_act in enumerate(manual_acts):
                if idx in matched_manual:
                    continue
                
                manual_time = self._extract_datetime(manual_act.get('date') or manual_act.get('started_at'))
                if not manual_time:
                    continue
                
                time_diff = abs((auto_time - manual_time).total_seconds() / 60)
                
                if time_diff <= self.time_tolerance.total_seconds() / 60:
                    if min_diff is None or time_diff < min_diff:
                        min_diff = time_diff
                        best_match = (idx, manual_act, time_diff)
            
            # Создаем результат сравнения
            if best_match:
                idx, manual_act, time_diff = best_match
                matched_manual.add(idx)
                
                # Сравниваем показатели
                count_diff = self._compare_counts(auto_act, manual_act)
                weight_diff = self._compare_weights(auto_act, manual_act)
                
                # Определяем тип совпадения
                if count_diff <= self.count_tolerance:
                    match_type = 'exact' if count_diff <= 5 else 'close'
                else:
                    match_type = 'mismatch'
                
                result = ComparisonResult(
                    auto_act=auto_act,
                    manual_act=manual_act,
                    match_type=match_type,
                    time_diff_minutes=time_diff,
                    count_diff_percent=count_diff,
                    weight_diff_percent=weight_diff
                )
            else:
                # Нет совпадения
                result = ComparisonResult(
                    auto_act=auto_act,
                    manual_act=None,
                    match_type='missing',
                    time_diff_minutes=float('inf'),
                    count_diff_percent=100.0
                )
            
            results.append(result)
        
        # Добавляем пропущенные ручные акты
        for idx, manual_act in enumerate(manual_acts):
            if idx not in matched_manual:
                result = ComparisonResult(
                    auto_act={},
                    manual_act=manual_act,
                    match_type='extra_manual',
                    time_diff_minutes=float('inf'),
                    count_diff_percent=100.0
                )
                results.append(result)
        
        self.comparisons = results
        logger.info(f"✅ Сопоставлено: {len(results)} пар")
        return results
    
    def calculate_metrics(self) -> Dict[str, float]:
        """
        Вычисляет метрики точности.
        
        Returns:
            Словарь с метриками: Recall, Precision, MAE, MAPE, Correlation
        """
        if not self.comparisons:
            return {}
        
        # Подсчет совпадений
        exact_matches = sum(1 for c in self.comparisons if c.match_type == 'exact')
        close_matches = sum(1 for c in self.comparisons if c.match_type == 'close')
        mismatches = sum(1 for c in self.comparisons if c.match_type == 'mismatch')
        missing = sum(1 for c in self.comparisons if c.match_type == 'missing')
        extra_manual = sum(1 for c in self.comparisons if c.match_type == 'extra_manual')
        
        total_auto = exact_matches + close_matches + mismatches + missing
        total_manual = exact_matches + close_matches + mismatches + extra_manual
        true_positives = exact_matches + close_matches
        
        # Recall (полнота)
        recall = true_positives / total_manual if total_manual > 0 else 0
        
        # Precision (точность)
        precision = true_positives / total_auto if total_auto > 0 else 0
        
        # F1-Score
        f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
        
        # MAE (Mean Absolute Error) по количеству
        count_errors = [
            abs(c.count_diff_percent)
            for c in self.comparisons
            if c.match_type in ['exact', 'close', 'mismatch']
        ]
        mae = sum(count_errors) / len(count_errors) if count_errors else 0
        
        # MAPE (Mean Absolute Percentage Error)
        mape = mae  # Уже в процентах
        
        # Корреляция (упрощенная)
        correlation = 1.0 - (mae / 100.0) if mae < 100 else 0
        
        metrics = {
            'recall': recall,
            'precision': precision,
            'f1_score': f1,
            'mae': mae,
            'mape': mape,
            'correlation': correlation,
            'exact_matches': exact_matches,
            'close_matches': close_matches,
            'mismatches': mismatches,
            'missing_auto': missing,
            'extra_manual': extra_manual,
            'total_auto': total_auto,
            'total_manual': total_manual
        }
        
        self.metrics = metrics
        logger.info(f"✅ Метрики вычислены: Recall={recall:.2%}, Precision={precision:.2%}")
        return metrics
    
    def generate_report(self, output_path: str) -> bool:
        """
        Генерирует Excel отчет о сверке.
        
        Создает файл с тремя листами:
        - Совпадения (зеленый)
        - Расхождения (желтый/красный)
        - Пропущенные (серый)
        """
        if not HAVE_OPENPYXL:
            logger.error("❌ Требуется openpyxl для генерации отчета")
            return False
        
        try:
            wb = Workbook()
            
            # Лист 1: Совпадения
            ws_matches = wb.active
            ws_matches.title = "Совпадения"
            self._fill_matches_sheet(ws_matches)
            
            # Лист 2: Расхождения
            ws_mismatches = wb.create_sheet("Расхождения")
            self._fill_mismatches_sheet(ws_mismatches)
            
            # Лист 3: Пропущенные
            ws_missing = wb.create_sheet("Пропущенные")
            self._fill_missing_sheet(ws_missing)
            
            # Лист 4: Метрики
            ws_metrics = wb.create_sheet("Метрики")
            self._fill_metrics_sheet(ws_metrics)
            
            wb.save(output_path)
            logger.info(f"✅ Отчет создан: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            return False
    
    def _fill_matches_sheet(self, ws):
        """Заполняет лист совпадений"""
        # Заголовки
        headers = ['Дата/Время', 'Авто (слева)', 'Авто (справа)', 'Ручн (слева)', 'Ручн (справа)', 'Разница %', 'Статус']
        ws.append(headers)
        
        # Стиль заголовка
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Данные
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        for comp in self.comparisons:
            if comp.match_type in ['exact', 'close']:
                row = [
                    comp.auto_act.get('started_at', ''),
                    comp.auto_act.get('left_count', 0),
                    comp.auto_act.get('right_count', 0),
                    comp.manual_act.get('left_count', 0) if comp.manual_act else 0,
                    comp.manual_act.get('right_count', 0) if comp.manual_act else 0,
                    f"{comp.count_diff_percent:.1f}%",
                    '✓ Совпадение'
                ]
                ws.append(row)
                
                # Зеленый фон
                for cell in ws[ws.max_row]:
                    cell.fill = green_fill
    
    def _fill_mismatches_sheet(self, ws):
        """Заполняет лист расхождений"""
        headers = ['Дата/Время', 'Авто', 'Ручн', 'Разница %', 'Статус']
        ws.append(headers)
        
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for comp in self.comparisons:
            if comp.match_type == 'mismatch':
                row = [
                    comp.auto_act.get('started_at', ''),
                    comp.auto_act.get('left_count', 0) + comp.auto_act.get('right_count', 0),
                    (comp.manual_act.get('left_count', 0) + comp.manual_act.get('right_count', 0)) if comp.manual_act else 0,
                    f"{comp.count_diff_percent:.1f}%",
                    '⚠ Расхождение'
                ]
                ws.append(row)
                
                # Желтый или красный фон
                fill = red_fill if comp.count_diff_percent > 20 else yellow_fill
                for cell in ws[ws.max_row]:
                    cell.fill = fill
    
    def _fill_missing_sheet(self, ws):
        """Заполняет лист пропущенных"""
        headers = ['Дата/Время', 'Тип', 'Количество', 'Примечание']
        ws.append(headers)
        
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        for comp in self.comparisons:
            if comp.match_type in ['missing', 'extra_manual']:
                if comp.match_type == 'missing':
                    row = [
                        comp.auto_act.get('started_at', ''),
                        'Авто (нет ручной)',
                        comp.auto_act.get('left_count', 0) + comp.auto_act.get('right_count', 0),
                        'Не найдено в ручных записях'
                    ]
                else:
                    row = [
                        comp.manual_act.get('date', '') if comp.manual_act else '',
                        'Ручн (нет авто)',
                        (comp.manual_act.get('left_count', 0) + comp.manual_act.get('right_count', 0)) if comp.manual_act else 0,
                        'Не обнаружено автоматически'
                    ]
                ws.append(row)
                
                for cell in ws[ws.max_row]:
                    cell.fill = gray_fill
    
    def _fill_metrics_sheet(self, ws):
        """Заполняет лист метрик"""
        if not self.metrics:
            self.calculate_metrics()
        
        ws.append(['Метрика', 'Значение'])
        ws.append(['Recall (Полнота)', f"{self.metrics.get('recall', 0):.2%}"])
        ws.append(['Precision (Точность)', f"{self.metrics.get('precision', 0):.2%}"])
        ws.append(['F1-Score', f"{self.metrics.get('f1_score', 0):.2%}"])
        ws.append(['MAE (Средняя ошибка)', f"{self.metrics.get('mae', 0):.1f}%"])
        ws.append(['MAPE', f"{self.metrics.get('mape', 0):.1f}%"])
        ws.append(['Корреляция', f"{self.metrics.get('correlation', 0):.2f}"])
        ws.append([])
        ws.append(['Точных совпадений', self.metrics.get('exact_matches', 0)])
        ws.append(['Близких совпадений', self.metrics.get('close_matches', 0)])
        ws.append(['Расхождений', self.metrics.get('mismatches', 0)])
        ws.append(['Пропущено (авто)', self.metrics.get('missing_auto', 0)])
        ws.append(['Лишних (ручн)', self.metrics.get('extra_manual', 0)])
    
    def _extract_datetime(self, value) -> Optional[datetime]:
        """Извлекает datetime из различных форматов"""
        if isinstance(value, datetime):
            return value
        elif isinstance(value, str):
            try:
                return datetime.fromisoformat(value.replace('Z', '+00:00'))
            except:
                return None
        return None
    
    def _compare_counts(self, auto_act: Dict, manual_act: Dict) -> float:
        """Сравнивает количество (возвращает разницу в %)"""
        auto_total = auto_act.get('left_count', 0) + auto_act.get('right_count', 0)
        manual_total = manual_act.get('left_count', 0) + manual_act.get('right_count', 0)
        
        if manual_total == 0:
            return 100.0 if auto_total > 0 else 0.0
        
        return abs(auto_total - manual_total) / manual_total * 100
    
    def _compare_weights(self, auto_act: Dict, manual_act: Dict) -> float:
        """Сравнивает вес (возвращает разницу в %)"""
        auto_weight = auto_act.get('total_weight', 0)
        manual_weight = manual_act.get('total_weight', 0)
        
        if manual_weight == 0:
            return 100.0 if auto_weight > 0 else 0.0
        
        return abs(auto_weight - manual_weight) / manual_weight * 100
