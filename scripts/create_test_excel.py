#!/usr/bin/env python3
"""
Создание тестового Excel файла с эталонными данными
для проверки тестового режима
"""

from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ Требуется openpyxl: pip install openpyxl")
    exit(1)

def create_test_excel():
    """Создает тестовый Excel файл"""
    
    # Создаем книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Взвешивания"
    
    # Заголовки
    headers = [
        "Дата",
        "Время начала",
        "Время окончания",
        "Длительность (сек)",
        "Проходы слева",
        "Проходы справа",
        "Всего проходов",
        "Пиковое количество",
        "Примечания"
    ]
    
    # Стиль заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Тестовые данные (примерные акты взвешивания)
    base_date = datetime.now().replace(hour=8, minute=0, second=0, microsecond=0)
    
    test_acts = [
        # (начало_мин, конец_мин, слева, справа, пик)
        (0, 2, 15, 14, 8),
        (10, 12, 20, 19, 10),
        (25, 27, 12, 11, 6),
        (40, 42, 18, 17, 9),
        (60, 63, 25, 24, 12),
    ]
    
    for row, (start_min, end_min, left, right, peak) in enumerate(test_acts, 2):
        start_time = base_date + timedelta(minutes=start_min)
        end_time = base_date + timedelta(minutes=end_min)
        duration = (end_time - start_time).total_seconds()
        
        ws.cell(row, 1, start_time.strftime("%Y-%m-%d"))
        ws.cell(row, 2, start_time.strftime("%H:%M:%S"))
        ws.cell(row, 3, end_time.strftime("%H:%M:%S"))
        ws.cell(row, 4, duration)
        ws.cell(row, 5, left)
        ws.cell(row, 6, right)
        ws.cell(row, 7, left + right)
        ws.cell(row, 8, peak)
        ws.cell(row, 9, f"Тестовый акт #{row-1}")
    
    # Автоширина столбцов
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = ws.column_dimensions[openpyxl.utils.get_column_letter(col)]
        
        for cell in ws[openpyxl.utils.get_column_letter(col)]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        column.width = adjusted_width
    
    # Сохраняем
    output_dir = Path('docs/examples')
    output_dir.mkdir(parents=True, exist_ok=True)
    
    output_path = output_dir / 'manual_records_test.xlsx'
    wb.save(output_path)
    
    print(f"✅ Тестовый Excel создан: {output_path}")
    print(f"   Записей: {len(test_acts)}")
    print(f"\n📋 Содержимое:")
    for i, (start_min, end_min, left, right, peak) in enumerate(test_acts, 1):
        print(f"   {i}. {start_min:02d}:{end_min:02d} мин, "
              f"слева={left}, справа={right}, пик={peak}")
    
    return output_path

if __name__ == "__main__":
    create_test_excel()
